"""Преобразование YAML в книгу XLSX с сохранением всех уровней вложенности.

Поддерживаются два макета листа данных.

Блочная таблица — для YAML вида «каталог записей», как nginx_http_upstreams:
словарь, где каждая запись — словарь полей. Каждая запись выводится блоком строк:
имя записи в первой колонке, скалярные поля и списки — по своим колонкам,
вложенные словари (main, backup, …) — группами колонок с двухуровневой шапкой.
Списки растягиваются вниз внутри блока, блоки разделяются пустой строкой.

Универсальная таблица — для любых других структур: каждое конечное значение
становится строкой, колонки «Уровень 1..N» показывают путь до значения
(ключи словарей и номера элементов списков «[1]», «[2]», …).
"""

import io
import re
from datetime import date, datetime, timezone

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_MIN_COLUMN_WIDTH = 10
_MAX_COLUMN_WIDTH = 60
_SPACER_WIDTH = 8

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_NAME_FONT = Font(bold=True)
_REPEAT_FONT = Font(color="A6A6A6")  # повтор родительского уровня — приглушённый серый
_LABEL_FONT = Font(bold=True)
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_SHEET_TITLE_FORBIDDEN_RE = re.compile(r"[\[\]:*?/\\]")

_MISSING = object()


class YamlConversionError(Exception):
    """Ошибка разбора YAML или построения XLSX."""


# --------------------------------------------------------------------------
# Разбор YAML
# --------------------------------------------------------------------------

def _load_all(yaml_text: str) -> list:
    return [doc for doc in yaml.safe_load_all(yaml_text) if doc is not None]


def _parse_documents(yaml_text: str) -> tuple[list, list[str]]:
    """Разбирает YAML; возвращает (документы, примечания о применённых исправлениях).

    Спецификация YAML запрещает символы табуляции, но в реальных файлах они
    встречаются внутри значений. Если строгий разбор не удался, пробуем ещё раз,
    заменив табуляции на пробелы.
    """
    notes = []
    try:
        documents = _load_all(yaml_text)
    except yaml.YAMLError as first_error:
        if "\t" not in yaml_text:
            raise YamlConversionError(f"Не удалось разобрать YAML: {first_error}") from first_error
        try:
            documents = _load_all(yaml_text.replace("\t", " "))
        except yaml.YAMLError:
            raise YamlConversionError(f"Не удалось разобрать YAML: {first_error}") from first_error
        notes.append(
            "В файле найдены символы табуляции — при разборе они заменены на пробелы."
        )
    if not documents:
        raise YamlConversionError("Файл не содержит данных YAML.")
    return documents, notes


# --------------------------------------------------------------------------
# Общие помощники
# --------------------------------------------------------------------------

def _is_scalar(value) -> bool:
    return value is None or isinstance(value, (str, int, float, bool, datetime, date))


def _is_scalar_list(value) -> bool:
    return isinstance(value, (list, tuple)) and all(_is_scalar(item) for item in value)


def _write_value_cell(cell, value):
    """Записывает значение с сохранением типа, безопасно для Excel."""
    if value is None:
        cell.value = "null"
    elif isinstance(value, bool) or isinstance(value, (int, float)):
        cell.value = value
    elif isinstance(value, datetime) and value.tzinfo is not None:
        cell.value = value.isoformat()  # Excel не поддерживает даты с таймзоной
    else:
        text = str(value)
        cell.value = text
        if text.startswith("="):
            cell.data_type = "s"  # не даём Excel принять строку за формулу


def _sheet_title(file_name: str) -> str:
    base = re.sub(r"\.(ya?ml)$", "", file_name, flags=re.IGNORECASE)
    title = _SHEET_TITLE_FORBIDDEN_RE.sub("_", base).strip()
    return title[:31] or "Данные"


# --------------------------------------------------------------------------
# Макет «блочная таблица» (каталог записей)
# --------------------------------------------------------------------------

def _record_fields(records: dict) -> list | None:
    """Описание полей записей в порядке первого появления или None, если не подходит.

    Поле может быть «значением» (скаляр или список скаляров — выводится одной
    колонкой, список растягивается вниз) либо «группой» (словарь из значений —
    выводится набором подколонок). Более глубокая вложенность в блочную таблицу
    не укладывается.
    """
    fields: dict[str, dict] = {}
    order: list[str] = []
    for record in records.values():
        if not isinstance(record, dict):
            return None
        for raw_key, value in record.items():
            key = str(raw_key)
            entry = fields.get(key)
            if entry is None:
                entry = {"kind": None, "sub": {}}
                fields[key] = entry
                order.append(key)
            if _is_scalar(value) or _is_scalar_list(value):
                if entry["kind"] == "group":
                    return None
                entry["kind"] = "value"
            elif isinstance(value, dict):
                if entry["kind"] == "value":
                    return None
                entry["kind"] = "group"
                for sub_key, sub_value in value.items():
                    if not (_is_scalar(sub_value) or _is_scalar_list(sub_value)):
                        return None
                    entry["sub"].setdefault(str(sub_key), None)
            else:
                return None
    return [
        {"key": key, "kind": fields[key]["kind"], "sub": list(fields[key]["sub"])}
        for key in order
    ]


def _extract_records(document):
    """(корневой ключ, записи, поля) для документа-«каталога записей», иначе None."""
    if not isinstance(document, dict) or len(document) != 1:
        return None
    root_key, inner = next(iter(document.items()))
    if not isinstance(inner, dict) or not inner:
        return None
    if not all(isinstance(value, dict) for value in inner.values()):
        return None
    fields = _record_fields(inner)
    if fields is None:
        return None
    return str(root_key), inner, fields


def _as_stack(value) -> list:
    """Значение поля как список ячеек, растягиваемых вниз внутри блока."""
    if value is _MISSING:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    return [value]


def _fill_record_sheet(sheet, records: dict, fields: list) -> int:
    # Физическая раскладка колонок: имя записи (без заголовка), поля-значения,
    # перед каждой группой — узкая колонка-разделитель с её названием.
    layout = [{"type": "name", "header": ""}]
    for field in fields:
        if field["kind"] == "value":
            layout.append({"type": "value", "key": field["key"], "header": field["key"]})
        else:
            layout.append(
                {"type": "spacer", "header": field["key"], "span": len(field["sub"])}
            )
            for sub in field["sub"]:
                layout.append(
                    {"type": "sub", "key": field["key"], "sub": sub, "header": sub}
                )
    total_columns = len(layout)

    # Шапка из двух строк: сначала стиль на все ячейки, затем объединения.
    for row in (1, 2):
        for column in range(1, total_columns + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER
    for column, spec in enumerate(layout, start=1):
        if spec["type"] in ("name", "value"):
            sheet.cell(row=1, column=column, value=spec["header"])
            sheet.merge_cells(
                start_row=1, start_column=column, end_row=2, end_column=column
            )
        elif spec["type"] == "spacer":
            sheet.cell(row=1, column=column, value=spec["header"])
            if spec["span"]:
                sheet.merge_cells(
                    start_row=1,
                    start_column=column,
                    end_row=1,
                    end_column=column + spec["span"],
                )
        else:
            sheet.cell(row=2, column=column, value=spec["header"])

    column_widths = [len(str(spec["header"] or "")) for spec in layout]

    row_cursor = 3
    for name, record in records.items():
        stacks = []
        for spec in layout:
            if spec["type"] == "value":
                stacks.append(_as_stack(record.get(spec["key"], _MISSING)))
            elif spec["type"] == "sub":
                group = record.get(spec["key"])
                sub_value = group.get(spec["sub"], _MISSING) if isinstance(group, dict) else _MISSING
                stacks.append(_as_stack(sub_value))
            else:
                stacks.append([])
        block_height = max([1] + [len(stack) for stack in stacks])

        name_cell = sheet.cell(row=row_cursor, column=1, value=str(name))
        name_cell.font = _NAME_FONT
        column_widths[0] = max(column_widths[0], len(str(name)))

        for column, (spec, stack) in enumerate(zip(layout, stacks), start=1):
            if spec["type"] == "spacer":
                continue
            for offset in range(block_height):
                cell = sheet.cell(row=row_cursor + offset, column=column)
                cell.border = _THIN_BORDER
                if offset < len(stack):
                    _write_value_cell(cell, stack[offset])
                    column_widths[column - 1] = max(
                        column_widths[column - 1], len(str(cell.value or ""))
                    )

        row_cursor += block_height + 1  # пустая строка-разделитель между блоками

    for column, spec in enumerate(layout, start=1):
        if spec["type"] == "spacer":
            width = _SPACER_WIDTH
        else:
            width = max(
                _MIN_COLUMN_WIDTH,
                min(_MAX_COLUMN_WIDTH, column_widths[column - 1] + 2),
            )
        sheet.column_dimensions[get_column_letter(column)].width = width

    sheet.freeze_panes = "A3"
    return len(records)


# --------------------------------------------------------------------------
# Макет «универсальная таблица» (уровни вложенности)
# --------------------------------------------------------------------------

def _walk(node, path, rows):
    """Рекурсивно собирает пары (путь, значение) для всех конечных значений."""
    if isinstance(node, dict):
        if not node:
            rows.append((path, "{}"))
        for key, value in node.items():
            _walk(value, path + (str(key),), rows)
    elif isinstance(node, (list, tuple, set)):
        if not node:
            rows.append((path, "[]"))
        for index, item in enumerate(node, start=1):
            _walk(item, path + (f"[{index}]",), rows)
    else:
        rows.append((path, node))


def _fill_levels_sheet(sheet, rows, depth):
    headers = [f"Уровень {level}" for level in range(1, depth + 1)] + ["Значение"]
    total_columns = len(headers)
    column_widths = [len(header) for header in headers]

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    previous_path = ()
    for row_index, (path, value) in enumerate(rows, start=2):
        same_prefix = True
        for level in range(depth):
            text = path[level] if level < len(path) else ""
            cell = sheet.cell(row=row_index, column=level + 1, value=text)
            cell.border = _THIN_BORDER
            is_repeat = (
                same_prefix
                and level < len(previous_path)
                and level < len(path)
                and path[level] == previous_path[level]
            )
            if is_repeat:
                cell.font = _REPEAT_FONT
            else:
                same_prefix = False
            if text:
                column_widths[level] = max(column_widths[level], len(text))

        value_cell = sheet.cell(row=row_index, column=total_columns)
        _write_value_cell(value_cell, value)
        value_cell.border = _THIN_BORDER
        column_widths[-1] = max(column_widths[-1], len(str(value_cell.value or "")))
        previous_path = path

    for column, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = max(
            _MIN_COLUMN_WIDTH, min(_MAX_COLUMN_WIDTH, width + 2)
        )

    last_row = len(rows) + 1
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(total_columns)}{last_row}"


# --------------------------------------------------------------------------
# Лист «Инфо» и сборка книги
# --------------------------------------------------------------------------

def _fill_info_sheet(sheet, *, source_url, file_name, stats, notes):
    info_rows = [
        ("Источник (URL)", source_url),
        ("Имя файла", file_name),
        ("Дата выгрузки (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
    ]
    info_rows.extend(stats)
    info_rows.extend(("Примечание", note) for note in notes)
    for row_index, (label, value) in enumerate(info_rows, start=1):
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        label_cell.font = _LABEL_FONT
        sheet.cell(row=row_index, column=2, value=value)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90


def convert_yaml_to_xlsx(
    yaml_text: str, *, source_url: str, file_name: str
) -> tuple[bytes, int, str | None]:
    """Строит книгу XLSX.

    Возвращает (байты файла, число записей/строк данных, корневой ключ документа —
    для блочной таблицы, иначе None).
    """
    documents, notes = _parse_documents(yaml_text)

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = _sheet_title(file_name)

    root_label = None
    extraction = _extract_records(documents[0]) if len(documents) == 1 else None
    if extraction is not None:
        root_label, records, fields = extraction
        count = _fill_record_sheet(data_sheet, records, fields)
        stats = [
            ("Макет", "блочная таблица"),
            ("Корневой ключ", root_label),
            ("Записей", count),
        ]
    else:
        rows = []
        if len(documents) == 1:
            _walk(documents[0], (), rows)
        else:
            for doc_index, document in enumerate(documents, start=1):
                _walk(document, (f"Документ {doc_index}",), rows)
        if not rows:
            raise YamlConversionError("В файле не найдено ни одного значения.")
        depth = max(len(path) for path, _ in rows)
        _fill_levels_sheet(data_sheet, rows, depth)
        count = len(rows)
        stats = [
            ("Макет", "универсальная таблица (уровни вложенности)"),
            ("Строк данных", count),
            ("Уровней вложенности", depth),
        ]

    info_sheet = workbook.create_sheet("Инфо")
    _fill_info_sheet(
        info_sheet,
        source_url=source_url,
        file_name=file_name,
        stats=stats,
        notes=notes,
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), count, root_label
